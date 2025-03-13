import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.cell.cell import Cell
from datetime import date
from openpyxl.utils import get_column_letter
import excel2img
from Account import Account
from mistake import Mistake

# Colors
yellow = "e1bc29"
red = "e15554"
blue = "4d9de0"
green = "3bb273"
purple = "7768ae"

accounts = []
mistakes = []
empty = [None, "", " "]

# Types of accounts
canadian_accounts = ["A", "C", "E", "G", "L", "Q", "S", "T", "X", "Y"]
us_accounts = ["B", "D", "F", "H", "M", "6", "R", "U"]

managed_prefixes = ["067", "087", "077", "057", "047"]
fee_based_prefixes = ["069", "089", "079", "029", "039", "059", "049"]

registered_suffixes = ["I", "J", "K", "O", "P", "Q", "6", "R", "S", "T", "U", "V", "W", "X", "Y", "Z"]
registered_sub_types = ["I", "J", "L", "C", "P", "T", "N", "X", "R", "S", "G", "D", "V-IND", "E-FAM", "U"]


def get_excel_column_name(column_number: int) -> str:
    column_number += 1
    column_name = ''

    while column_number > 0:
        modulo = (column_number - 1) % 26
        column_name = chr(ord('A') + modulo) + column_name
        column_number = (column_number - modulo) // 26
    print(column_number, " ", column_name)
    return column_name


# Returns the unique rows of a spreadsheet in a list
def get_unique_rows(spreadsheet) -> []:
    unique_rows = []
    for index, row in enumerate(spreadsheet.iter_rows()):
        sublist = []
        for cell in row:
            sublist.append(cell.value)

        if sublist not in unique_rows:
            unique_rows.append(sublist)
    return unique_rows


def color_cells(*cells, color: str):
    for cell in cells:
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")


def check_for_incorrect_funds(account_number: Cell, fund: Cell) -> bool:
    last_digit = account_number.value[-1]
    if (last_digit in canadian_accounts and fund.value != "C") \
            or (last_digit in us_accounts and fund.value != "U"):
        return True
    return False


def check_for_empty_cells(cell: Cell) -> bool:
    if cell.value in empty:
        return True
    return False


def managed_account(row, account_column) -> bool:
    account_prefix = row[account_column].value[:3]
    return account_prefix in managed_prefixes


def fee_based_account(row, account_column) -> bool:
    account_prefix = row[account_column].value[:3]
    return account_prefix in fee_based_prefixes


def check_for_port_flag(row, account_number, port_flag) -> bool:
    if managed_account(row, account_number) and row[port_flag].value is not "M":
        return True
    if fee_based_account(row, account_number) and row[port_flag].value is not "F":
        return True
    # We have a port flag, but our account isn't managed or fee based
    if row[port_flag].value not in empty and not (managed_account(row, account_number) or fee_based_account(row, account_number)):
        return True
    return False



# Load the spreadsheets
workbook = load_workbook(filename="04122023.xlsx")
new_workbook = openpyxl.Workbook()

raw = workbook.active
edited = new_workbook.active
edited.title = "Data"




# Remove duplicate Rows
unique_rows = get_unique_rows(raw)

# Add the unique rows to the edited sheet
for row in unique_rows:
    edited.append(row)


for row in edited.iter_rows(2):
    accounts.append(Account(edited, row))

for account in accounts:
    print(account.number_prefix, " ", account.port_type, " ",account.is_managed, " ", account.is_fee_based, " ", account.is_commission)

# Begin checking for errors and coloring the cells
# Outer loop goes from row to row
# for index, row in enumerate(edited.iter_rows(2)):
#     # Call the check for incorrect funds function on each row
#     if check_for_incorrect_funds(row[acct_number], row[funds]):
#         mistakes.append(Mistake(row, [row[acct_number], row[funds]], blue))

#     # Check port flag for managed and fee based accounts on each row
#     if check_for_port_flag(row, acct_number, port_type):
#         mistakes.append((Mistake(row, [row[acct_number], row[port_type]], yellow)))

#     # Inner loop goes cell to cell
#     for column_number, cell in enumerate(row):
#         # Call the check for empty cells function on each cell
#         if column_number in mandatory_columns and check_for_empty_cells(cell):
#             # Continue over this case because UCI is allowed to be blank here
#             if column_number is uci and \
#                     row[recip_type].value in [3, '3'] and \
#                     row[title].value in ["BENEFICIAL OWNER", "TA", "T/A", "PA",
#                                          "P/A", "BENEFICIAL OWNER / TA", "BENE OWNER / T/A",
#                                          "BENEFICIAL OWNER / T/A"]:
#                 # This is not a mistake so its colored green
#                 mistakes.append(Mistake(row, [cell, row[recip_type], row[title]], green))
#                 break
#             mistakes.append(Mistake(row, [cell], red))

# for mistake in mistakes:
#     mistake.show()

# # Adjust column widths to text length
# dims = {}
# for row in edited.rows:
#     for cell in row:
#         if cell.value:
#             dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
# for col, value in dims.items():
#     edited.column_dimensions[col].width = value

# Name and save the file
today = date.today()
name = '{:0>2}-{:0>2}-{}.xlsx'.format(str(today.month), str(today.day), str(today.year)[2:])
new_workbook.save(filename=name)

# for mistake in mistakes:
#     print(mistake)
#     count += 1
#     cols = get_excel_column_name(mistake[0]) + str(mistake[1]) + ":" + get_excel_column_name(mistake[2]) + str(
#         mistake[3])
#     excel2img.export_img("Mistakes.xlsx", "image" + str(count) + ".png", "test", cols)


